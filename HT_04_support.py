#! /usr/bin/env python
#  -*- coding: utf-8 -*-
#
# Support module generated by PAGE version 7.3
#  in conjunction with Tcl version 8.6
#    Feb 19, 2023 05:43:37 PM IST  platform: Windows NT
#    May 14, 2024 12:35:47 PM IST  platform: Windows NT

import sys
import tkinter as tk
import tkinter.ttk as ttk
from tkinter.constants import *
import serial,time
from serial import Serial
import serial.tools.list_ports
import openpyxl

import threading
import time

import HT_04

ser = serial.Serial()
wb = openpyxl.Workbook() 

sheet = wb.active
Connect_flag = 0

def main(*args):
    '''Main entry point for the application.'''
    global root
    root = tk.Tk()
    root.protocol( 'WM_DELETE_WINDOW' , root.destroy)
    # Creates a toplevel widget.
    global _top1, _w1
    _top1 = root
    _w1 = HT_04.Toplevel1(_top1)

    # Start the thread for updating the Text widget
    # update_thread = threading.Thread(target=update_text_widget)
    # update_thread.daemon = True  # Daemonize the thread to stop it when the main program exits
    # update_thread.start()

    #radio = IntVar()
    init_fun()
    refresh_ports()
    _w1.Text1.delete("1.0", END) 
    _w1.Text1.insert("1.0", 'Hello ^-^')

    # Start the thread for updating the Scrolledwindow1 widget
    update_thread = threading.Thread(target=update_scrolled_window)
    update_thread.daemon = True  # Daemonize the thread to stop it when the main program exits
    update_thread.start()
    
    root.mainloop()
    #_w1.Text1.delete("1.0", END) 


def update_scrolled_window():
    print("Updating Scrolledwindow1...")
    while True:
        # Generate or fetch new data
        new_data = "New data here..."
            
        # Update the Scrolledwindow1 widget with the new data
        _w1.Scrolledwindow1_f.configure(text=new_data)
            
        # Scroll to the end of the Scrolledwindow1 widget to show the new data
        _w1.Scrolledwindow1_f.yview_moveto(1.0)
            
        # Sleep for some time before fetching new data again
        time.sleep(1)  # Adjust the sleep time as needed


def init_fun(*args):
    global run_flag, max_flag

    #max_flag = 0
    run_flag = 0
    _w1.Text1.delete("1.0", END) 
    _w1.Text1.insert("1.0", 'Setting UP')
    
def Connect(*args):
    
    p = _w1.TCombobox1.get()
    p = p.split(" -")
    N = p[0]

    global Connect_flag,Test_flag

    
     

    if Connect_flag == 0:
    
        ser.port = N
        ser.baudrate = 115200
        ser.bytesize = 8
        ser.timeout = 2
        ser.parity = serial.PARITY_NONE
        ser.stopbits = serial.STOPBITS_ONE
        ser.open()
        if ser.is_open:
            _w1.Text1.delete("1.0", END) 
            _w1.Text1.insert(END, 'Port Cold not open')

        #p_x2 = "open"
        _w1.Button1.configure(background = "#50ef21")
        _w1.Button1.configure(text='''Connected''')
        Connect_flag = 1
        Test_flag = 1
        _w1.Text1.delete("1.0", END) 
        _w1.Text1.insert(END, 'Connected to entered port')
    else:
        Disconnect()
        Connect_flag = 0
        Test_flag = 0
        

def Disconnect(*args):
    ser.close()
    _w1.Button1.configure(background = "#d9d9d9")
    _w1.Button1.configure(state = NORMAL)
    _w1.Button1.configure(text='''Disconnected''')
    p_x2 = "not open"
    _w1.Text1.delete("1.0", END) 
    _w1.Text1.insert(END, 'Port is disconnected')

def show_ports():
    return serial.tools.list_ports.comports()

def refresh_ports(*args):
    _w1.TCombobox1['values'] = show_ports()
    time.sleep(1)
    _w1.Text1.delete("1.0", END) 
    _w1.Text1.insert(END, 'Hello again ^-^')
    

def run(*args):

    max_flag = max_pin()
    file_flag = file_name()

    run_flag = max_flag and file_flag
    

    if run_flag != 1:
        if max_flag == 0:
           _w1.Text1.delete("1.0", END) 
           _w1.Text1.insert(END, 'Enter valid pin number')
        else:
            _w1.Text1.delete("1.0", END) 
            _w1.Text1.insert(END, 'Enter file name x_x')
        return 0

    close_flag = 0 
    i = 0
    x = 0

    c1 = sheet.cell(row = 1, column = 5) 
            # writing values to cells 
    c1.value = "Test Report"

    c2 = sheet.cell(row = 2, column = 1) 
            # writing values to cells 
    c2.value = "Side 1"

    
    c3 = sheet.cell(row = 2, column = 2) 
            # writing values to cells 
    c3.value = "Side 2"

    check_port()

    test = 13
    
    ser.write(1)
    run_flag = 0
    max_flag = 0
    file_flag = 0

    while i != 255:

        i = 0
        i = int.from_bytes(ser.read(1), "big")
        if x > max_p + 1:
            wb.save(f_name)
            #x = 0
            _w1.Button3.configure(background="#d9d9d9")
            return 1

        if i == 255:
            wb.save(f_name)
            x = 0
            _w1.Button3.configure(background="#d9d9d9")
            return 1
        
        else:
            if i == 254:
                c1 = sheet.cell(row = x + 3, column = 1) 
                c1.value = x
                c1 = sheet.cell(row = x + 3, column = 2) 
                c1.value = "NC"
            
            else:
                c1 = sheet.cell(row = x + 3, column = 1) 
                c1.value = x
                c1 = sheet.cell(row = x + 3, column = 2) 
                c1.value = i
        
            if i == 0:

                close_flag += 1

                if close_flag > 1:
                    _w1.Button3.configure(background = "#FF0000")
                    _w1.Button3.configure(text = '''Connect failed''')
                    #self.Button3.configure(background="#d9d9d9")
                    time.sleep(1)
                    _w1.Button3.configure(text = '''Run''')
                    #_w1.Text1.configure(text = '''Communication Failed''')
                    _w1.Text1.delete("1.0", END) 
                    _w1.Text1.insert(END, 'Communication Failed')
                    return 1
        x += 1
    #wb.save(f_name) 
   
#add file section working without concern
        
def xcl_format(*args):
    c1 = sheet.cell(row = 1, column = 5) 
            # writing values to cells 
    c1.value = "Test Report"

    c2 = sheet.cell(row = 2, column = 1) 
            # writing values to cells 
    c2.value = "Side 1"

    
    c3 = sheet.cell(row = 2, column = 2) 
            # writing values to cells 
    c3.value = "Side 2"

    for x in range(50):
            c1 = sheet.cell(row = x + 3, column = 1) 
            c1.value = x
            # writing values to cells 
            #c1.value = ser.read(1).hex()
            

    wb.save(f_name)

def file_name(*args):
    global f_name,file_flag
    file_flag = 1
    f_name = 'default.xlsx'
    f_name = _w1.Entry1.get()
    if f_name == '':
        _w1.Text1.delete("1.0", END) 
        _w1.Text1.insert(END, 'Please enter file name x_x')
        return 0
        
    f_name += ".xlsx"
    _w1.Text1.delete("1.0", END) 
    _w1.Text1.insert(END, 'File name is '+ f_name)
    return 1

def max_pin(*args):
    global max_p
    max_flag = 0
    max_p = str(_w1.Entry2.get())
    
    if max_p == '':
        max_p = '0'

    max_p = int(max_p)

    if max_p < 50 and max_p > 0:
        _w1.Text1.delete("1.0", END) 
        _w1.Text1.insert(END, 'Got it')
        max_flag = 1
        

    else:
        _w1.Text1.delete("1.0", END) 
        _w1.Text1.insert(END, 'Enter valid pin number, should be between 0 to 50')
        max_flag = 0
    return max_flag
        

def check_port(*args):
    if(ser.isOpen() == False):
        _w1.Text1.delete("1.0", END) 
        _w1.Text1.insert(END, 'Port is not open')
    #ser.open()

def Test(*args):
    # if _debug:
    print('HT_04_support.Test')
    for arg in args:
        print ('    another arg:', arg)
    sys.stdout.flush()

if __name__ == '__main__':
    HT_04.start_up()





